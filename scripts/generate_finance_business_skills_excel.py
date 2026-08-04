#!/usr/bin/env python3
"""Build the finance business Skill evaluation workbook from a completed JSON run."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from statistics import mean, median
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17365D"
BLUE = "2563EB"
LIGHT_BLUE = "EAF2FF"
PALE_BLUE = "F5F8FF"
GREEN = "12805C"
LIGHT_GREEN = "E9F8F1"
ORANGE = "B45309"
LIGHT_ORANGE = "FFF4E5"
RED = "B42318"
LIGHT_RED = "FDECEC"
GRAY = "667085"
LIGHT_GRAY = "F2F4F7"
WHITE = "FFFFFF"
BORDER_COLOR = "D0D5DD"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


REVIEWS = {
    "market_overview_real": (
        "基本可用",
        "入口、数据查询和最终结论完整；指数、成交与风格分化均有明确数据支撑。",
    ),
    "sector_theme_real": (
        "未完成",
        "Skill 入口正确并取得白酒行业成分，但 28 次工具调用后仍未形成最终回答。",
    ),
    "stock_research_real": (
        "部分可用",
        "研究内容和证据较丰富，但最终回答达到 2000 字上限后被截断，风险与观察变量不完整。",
    ),
    "earnings_analysis_real": (
        "基本可用",
        "能够围绕增长质量形成结论并说明现金流数据缺口；部分证据来自研报搜索。",
    ),
    "stock_screening_real": (
        "部分可用",
        "筛选逻辑可执行且披露了数据限制，但仅覆盖 100 只成分样本，ROE 只有少量记录。",
    ),
    "factor_analysis_real": (
        "部分可用",
        "完成动量与波动率横截面分析，但样本覆盖有限且答案被 2000 字上限截断。",
    ),
    "valuation_analysis_real": (
        "部分可用",
        "估值位置判断清楚，但近三年盈利、ROE 和现金流数据缺失，只能给出条件性结论。",
    ),
    "financial_quality_real": (
        "数据不足",
        "Skill 正确识别并拒绝编造，但目标公司的三表数据为零行，无法完成财务质量判断。",
    ),
    "stock_comparison_real": (
        "部分可用",
        "行情与估值比较有效，但成长、盈利、现金流和资产负债证据缺失，且答案被截断。",
    ),
    "technical_structure_real": (
        "基本可用",
        "趋势、均线、成交、波动率、支撑阻力和失效条件均有清晰数据与结论。",
    ),
    "dividend_analysis_real": (
        "部分可用",
        "分红记录与股息率分析较充分，但缺少利润和现金流数据，持续性判断仍是条件性的。",
    ),
}


def _safe(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _set_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(2, 1, subtitle)
    cell.font = Font(name="Aptos", size=10, color=GRAY)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def _style_header(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def _style_body(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=WHITE if row % 2 else PALE_BLUE)
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Aptos", size=9, color="101828")
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_widths(ws, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _status_fill(value: str) -> PatternFill:
    if value in {"基本可用", "完成", "是"}:
        return PatternFill("solid", fgColor=LIGHT_GREEN)
    if value in {"部分可用", "数据不足"}:
        return PatternFill("solid", fgColor=LIGHT_ORANGE)
    if value in {"未完成", "否"}:
        return PatternFill("solid", fgColor=LIGHT_RED)
    return PatternFill("solid", fgColor=LIGHT_GRAY)


def build_workbook(report: dict[str, Any], output_path: Path) -> None:
    cases = report["cases"]
    durations = [case["total_duration_ms"] / 1000 for case in cases]
    total_calls = sum(case.get("tool_call_count", 0) for case in cases)
    completed = sum(bool(case.get("execution_completed")) for case in cases)
    matched = sum(bool(case.get("skill_match")) for case in cases)
    truncated = sum(bool(case.get("answer_maybe_truncated")) for case in cases)

    wb = Workbook()
    ws = wb.active
    ws.title = "评测总览"
    _set_title(
        ws,
        "金融业务 Skills 真实执行评测",
        f"运行时间：{report.get('started_at', '')} 至 {report.get('finished_at', '')}｜"
        f"运行模式：{report.get('runtime_mode', '')}",
        12,
    )

    kpis = [
        ("评测 Skill", len(cases), "个"),
        ("入口匹配", matched, f"/ {len(cases)}"),
        ("形成回答", completed, f"/ {len(cases)}"),
        ("总耗时", round(sum(durations), 1), "秒"),
        ("平均耗时", round(mean(durations), 1), "秒/例"),
        ("工具调用", total_calls, "次"),
        ("答案截断", truncated, "例"),
    ]
    for idx, (label, value, unit) in enumerate(kpis):
        col = 1 + idx
        ws.cell(4, col, label)
        ws.cell(5, col, f"{value} {unit}")
        ws.cell(4, col).font = Font(name="Aptos", size=9, bold=True, color=GRAY)
        ws.cell(5, col).font = Font(name="Aptos Display", size=16, bold=True, color=NAVY)
        for row in (4, 5):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, col).border = THIN_BORDER
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 32

    headers = [
        "序号",
        "Case ID",
        "预期 Skill",
        "实际 Skill",
        "入口匹配",
        "执行结果",
        "业务评审",
        "总耗时(秒)",
        "首次进展(秒)",
        "工具调用",
        "步骤数",
        "评审摘要",
    ]
    header_row = 8
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
    _style_header(ws, header_row, 1, len(headers))

    for idx, case in enumerate(cases, 1):
        row = header_row + idx
        review, review_note = REVIEWS.get(case["id"], ("待复核", ""))
        values = [
            idx,
            case["id"],
            case.get("expected_skill_id", ""),
            ", ".join(case.get("actual_skill_ids") or []),
            "是" if case.get("skill_match") else "否",
            "完成" if case.get("execution_completed") else "未完成",
            review,
            case.get("total_duration_ms", 0) / 1000,
            case.get("first_progress_ms", 0) / 1000,
            case.get("tool_call_count", 0),
            case.get("step_count", 0),
            review_note,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.row_dimensions[row].height = 58
    _style_body(ws, header_row + 1, header_row + len(cases), 1, len(headers))
    for row in range(header_row + 1, header_row + len(cases) + 1):
        for col in (5, 6, 7):
            ws.cell(row, col).fill = _status_fill(_safe(ws.cell(row, col).value))
            ws.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color="101828")
        for col in (1, 5, 6, 7, 8, 9, 10, 11):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 8).number_format = "0.0"
        ws.cell(row, 9).number_format = "0.0"
    ws.auto_filter.ref = f"A{header_row}:L{header_row + len(cases)}"
    ws.freeze_panes = f"A{header_row + 1}"
    _set_widths(
        ws,
        {
            1: 7,
            2: 27,
            3: 24,
            4: 24,
            5: 11,
            6: 12,
            7: 13,
            8: 13,
            9: 14,
            10: 12,
            11: 10,
            12: 58,
        },
    )

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "各 Skill 总耗时"
    chart.y_axis.title = "Skill"
    chart.x_axis.title = "秒"
    data = Reference(ws, min_col=8, min_row=header_row, max_row=header_row + len(cases))
    categories = Reference(
        ws, min_col=3, min_row=header_row + 1, max_row=header_row + len(cases)
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7.5
    chart.width = 15
    ws.add_chart(chart, "A22")

    details = wb.create_sheet("案例详情")
    _set_title(
        details,
        "案例详情与最终结果",
        "每行对应一个真实问题；业务评审由执行结果、证据覆盖和回答完整性综合判断。",
        13,
    )
    detail_headers = [
        "序号",
        "Case ID",
        "用户问题",
        "预期 Skill",
        "实际 Skill",
        "路由",
        "执行完成",
        "业务评审",
        "评审摘要",
        "总耗时(秒)",
        "答案字符数",
        "是否截断",
        "最终回答 / 错误现场",
    ]
    for col, header in enumerate(detail_headers, 1):
        details.cell(4, col, header)
    _style_header(details, 4, 1, len(detail_headers))
    for idx, case in enumerate(cases, 1):
        row = 4 + idx
        review, review_note = REVIEWS.get(case["id"], ("待复核", ""))
        route = case.get("routing") or {}
        answer = case.get("answer") or case.get("error") or ""
        values = [
            idx,
            case["id"],
            case.get("question", ""),
            case.get("expected_skill_id", ""),
            ", ".join(case.get("actual_skill_ids") or []),
            f"{route.get('selected_agent', '')} / {route.get('turn_mode', '')}",
            "是" if case.get("execution_completed") else "否",
            review,
            review_note,
            case.get("total_duration_ms", 0) / 1000,
            case.get("answer_char_count", 0),
            "是" if case.get("answer_maybe_truncated") else "否",
            answer,
        ]
        for col, value in enumerate(values, 1):
            details.cell(row, col, value)
        details.row_dimensions[row].height = 180
    _style_body(details, 5, 4 + len(cases), 1, len(detail_headers))
    for row in range(5, 5 + len(cases)):
        details.cell(row, 7).fill = _status_fill(details.cell(row, 7).value)
        details.cell(row, 8).fill = _status_fill(details.cell(row, 8).value)
        details.cell(row, 12).fill = _status_fill("未完成" if details.cell(row, 12).value == "是" else "")
        details.cell(row, 10).number_format = "0.0"
    details.auto_filter.ref = f"A4:M{4 + len(cases)}"
    details.freeze_panes = "A5"
    _set_widths(
        details,
        {
            1: 7,
            2: 26,
            3: 55,
            4: 23,
            5: 23,
            6: 27,
            7: 12,
            8: 13,
            9: 55,
            10: 13,
            11: 13,
            12: 11,
            13: 95,
        },
    )

    tool_sheet = wb.create_sheet("工具调用")
    _set_title(
        tool_sheet,
        "工具调用明细",
        "记录每个案例调用的系统工具、金融主体和数据视图；同一工具可能多次调用。",
        7,
    )
    tool_headers = ["Case ID", "序号", "工具", "Subject", "Dataview", "调用时点/来源", "备注"]
    for col, header in enumerate(tool_headers, 1):
        tool_sheet.cell(4, col, header)
    _style_header(tool_sheet, 4, 1, len(tool_headers))
    row = 5
    for case in cases:
        for idx, call in enumerate(case.get("tool_calls") or [], 1):
            tool_sheet.append(
                [
                    case["id"],
                    idx,
                    call.get("tool", ""),
                    call.get("subject", ""),
                    call.get("dataview", ""),
                    call.get("elapsed_ms", ""),
                    _safe({k: v for k, v in call.items() if k not in {"tool", "subject", "dataview", "elapsed_ms"}}),
                ]
            )
            row += 1
    _style_body(tool_sheet, 5, row - 1, 1, len(tool_headers))
    tool_sheet.auto_filter.ref = f"A4:G{row - 1}"
    tool_sheet.freeze_panes = "A5"
    _set_widths(tool_sheet, {1: 28, 2: 9, 3: 29, 4: 18, 5: 24, 6: 16, 7: 70})

    step_sheet = wb.create_sheet("执行步骤")
    _set_title(
        step_sheet,
        "执行步骤与进展",
        "步骤来自真实流式事件，保留阶段、耗时、状态和面向用户的过程说明。",
        9,
    )
    step_headers = [
        "Case ID",
        "序号",
        "相对耗时(秒)",
        "阶段",
        "标题",
        "事件类型",
        "状态",
        "内容",
        "对应 Skill",
    ]
    for col, header in enumerate(step_headers, 1):
        step_sheet.cell(4, col, header)
    _style_header(step_sheet, 4, 1, len(step_headers))
    row = 5
    for case in cases:
        skill_name = ", ".join(case.get("actual_skill_ids") or [])
        for idx, step in enumerate(case.get("steps") or [], 1):
            step_sheet.append(
                [
                    case["id"],
                    idx,
                    step.get("elapsed_ms", 0) / 1000,
                    step.get("stage", ""),
                    step.get("title", ""),
                    step.get("type", ""),
                    step.get("status", ""),
                    step.get("content", ""),
                    skill_name,
                ]
            )
            row += 1
    _style_body(step_sheet, 5, row - 1, 1, len(step_headers))
    for body_row in range(5, row):
        step_sheet.cell(body_row, 3).number_format = "0.000"
        status = _safe(step_sheet.cell(body_row, 7).value)
        if status == "error":
            step_sheet.cell(body_row, 7).fill = PatternFill("solid", fgColor=LIGHT_RED)
        elif status == "completed":
            step_sheet.cell(body_row, 7).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    step_sheet.auto_filter.ref = f"A4:I{row - 1}"
    step_sheet.freeze_panes = "A5"
    _set_widths(
        step_sheet,
        {1: 28, 2: 9, 3: 15, 4: 15, 5: 25, 6: 28, 7: 13, 8: 80, 9: 25},
    )

    request_sheet = wb.create_sheet("数据请求")
    _set_title(
        request_sheet,
        "金融数据请求与结果引用",
        "记录 CC 实际发出的金融查询、结果引用和返回规模，便于审计数据取得过程。",
        10,
    )
    request_headers = [
        "Case ID",
        "序号",
        "业务目标",
        "API",
        "请求",
        "结果变量",
        "结果引用",
        "记录数",
        "数据类型",
        "Flow Step",
    ]
    for col, header in enumerate(request_headers, 1):
        request_sheet.cell(4, col, header)
    _style_header(request_sheet, 4, 1, len(request_headers))
    row = 5
    for case in cases:
        for idx, ref in enumerate(case.get("result_refs") or [], 1):
            request_sheet.append(
                [
                    case["id"],
                    idx,
                    ref.get("goal", ""),
                    ref.get("api", ""),
                    ref.get("request", ""),
                    ref.get("result_name", ""),
                    ref.get("result_ref", ""),
                    ref.get("row_count", ""),
                    ref.get("data_type", ""),
                    ref.get("flow_step", ""),
                ]
            )
            row += 1
    _style_body(request_sheet, 5, row - 1, 1, len(request_headers))
    request_sheet.auto_filter.ref = f"A4:J{row - 1}"
    request_sheet.freeze_panes = "A5"
    _set_widths(
        request_sheet,
        {1: 28, 2: 9, 3: 45, 4: 24, 5: 90, 6: 15, 7: 60, 8: 12, 9: 16, 10: 12},
    )

    metadata = wb.create_sheet("运行说明")
    _set_title(metadata, "运行说明", "工作簿来源、口径和关键限制。", 4)
    notes = [
        ("原始记录", str(report.get("_source_path", ""))),
        ("评测集版本", report.get("version", "")),
        ("运行模式", report.get("runtime_mode", "")),
        ("最大 Agent Turns", report.get("max_turns", "")),
        ("案例数量", len(cases)),
        ("总耗时（秒）", round(sum(durations), 1)),
        ("耗时中位数（秒）", round(median(durations), 1)),
        ("入口匹配", f"{matched}/{len(cases)}"),
        ("形成最终回答", f"{completed}/{len(cases)}"),
        ("工具调用总数", total_calls),
        ("回答截断", f"{truncated} 个案例达到系统 2000 字保存上限"),
        (
            "评审口径",
            "入口正确、技术执行完成、业务结果可用是三个不同维度；数据不足但如实说明不等于 Skill 入口失败。",
        ),
        (
            "生成时间",
            datetime.now().astimezone().isoformat(timespec="seconds"),
        ),
    ]
    metadata.cell(4, 1, "项目")
    metadata.cell(4, 2, "内容")
    _style_header(metadata, 4, 1, 2)
    for idx, (label, value) in enumerate(notes, 5):
        metadata.cell(idx, 1, label)
        metadata.cell(idx, 2, value)
        metadata.row_dimensions[idx].height = 36
    _style_body(metadata, 5, 4 + len(notes), 1, 2)
    metadata.freeze_panes = "A5"
    _set_widths(metadata, {1: 28, 2: 110})

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.outlinePr.summaryBelow = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def verify_workbook(output_path: Path, expected_cases: int) -> dict[str, Any]:
    wb = load_workbook(output_path, data_only=False, read_only=False)
    required = ["评测总览", "案例详情", "工具调用", "执行步骤", "数据请求", "运行说明"]
    missing = [name for name in required if name not in wb.sheetnames]
    formula_errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(
                    token in value
                    for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}:{value}")
    details_count = wb["案例详情"].max_row - 4
    return {
        "sheet_names": wb.sheetnames,
        "missing_sheets": missing,
        "details_count": details_count,
        "expected_cases": expected_cases,
        "formula_errors": formula_errors,
        "file_size": output_path.stat().st_size,
        "valid": not missing and details_count == expected_cases and not formula_errors,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    with args.input.open("r", encoding="utf-8") as handle:
        report = json.load(handle)
    report["_source_path"] = str(args.input.resolve())
    build_workbook(report, args.output)
    verification = verify_workbook(args.output, len(report["cases"]))
    print(json.dumps(verification, ensure_ascii=False, indent=2))
    if not verification["valid"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
