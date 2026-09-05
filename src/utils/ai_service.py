import argparse
import os ,re , json, base64, mimetypes
import openpyxl
from openai import OpenAI
from fastapi import FastAPI, HTTPException
from starlette.responses import StreamingResponse
import asyncio
from pathlib import Path
from dotenv import load_dotenv


# Load the repository configuration before constructing any provider client.
# Flask used to do this in its entrypoint, which made standalone services and
# eval scripts silently fall back to ``not-configured`` credentials.
REPO_ROOT = Path(__file__).resolve().parents[2]
load_dotenv(REPO_ROOT / ".env", override=False)

client_alibaba = OpenAI(
    api_key=os.getenv("DASHSCOPE_API_KEY") or "not-configured",
    base_url=os.getenv("DASHSCOPE_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1"),
    timeout=float(os.getenv("LLM_CLIENT_TIMEOUT_SECONDS", "45")),
    max_retries=int(os.getenv("LLM_CLIENT_MAX_RETRIES", "1")),
)   



client = OpenAI(
    api_key=os.getenv("OPENAI_PROXY_API_KEY") or "not-configured",
    #base_url="http://ec2-54-253-120-126.ap-southeast-2.compute.amazonaws.com:8080/v1/",
    base_url=os.getenv("OPENAI_PROXY_BASE_URL", "http://api.openai-proxy.com/v1/"),
    timeout=float(os.getenv("LLM_CLIENT_TIMEOUT_SECONDS", "45")),
    max_retries=int(os.getenv("LLM_CLIENT_MAX_RETRIES", "1")),
)

client_next = OpenAI(
    api_key=os.getenv("NEXT_API_KEY") or "not-configured",
    #base_url="http://ec2-54-253-120-126.ap-southeast-2.compute.amazonaws.com:8080/v1/",
    base_url=os.getenv("NEXT_API_BASE_URL", "https://api.nextapi.fun/v1/"),
    timeout=float(os.getenv("LLM_CLIENT_TIMEOUT_SECONDS", "45")),
    max_retries=int(os.getenv("LLM_CLIENT_MAX_RETRIES", "1")),
)


llm_client = OpenAI(
    api_key=(
        os.getenv("LLM_API_KEY")
        or os.getenv("LLM_KEY")
        or os.getenv("DEEPSEEK_API_KEY")
        or os.getenv("DASHSCOPE_API_KEY")
        or "not-configured"
    ),
    base_url=(
        os.getenv("LLM_BASE_URL")
        or os.getenv("LLM_ENDPOINT")
        or os.getenv("DASHSCOPE_BASE_URL")
        or "https://api.deepseek.com/v1"
    ),
    timeout=float(os.getenv("LLM_CLIENT_TIMEOUT_SECONDS", "45")),
    max_retries=int(os.getenv("LLM_CLIENT_MAX_RETRIES", "1")),
)

# Backward-compatible alias. Old call sites still reference a "deepseek" client.
deepseek_client = llm_client

DEFAULT_CHAT_MODEL = os.getenv("LLM_DEFAULT_MODEL", "deepseek-chat")
DEFAULT_FLASH_MODEL = os.getenv("LLM_FLASH_MODEL", "deepseek-chat")
DEFAULT_REASONING_MODEL = os.getenv("LLM_REASONING_MODEL", DEFAULT_CHAT_MODEL)
DEFAULT_EMBEDDING_MODEL = os.getenv("LLM_EMBEDDING_MODEL", "text-embedding-v4")
DEFAULT_ENABLE_THINKING = os.getenv("LLM_DEFAULT_ENABLE_THINKING", "false").lower() == "true"
DEFAULT_MAX_TOKENS = int(os.getenv("LLM_DEFAULT_MAX_TOKENS", "8192"))
DEFAULT_LONG_CONTEXT_MAX_TOKENS = int(os.getenv("LLM_LONG_CONTEXT_MAX_TOKENS", "40960"))


def llm_config_summary() -> dict:
    """Return safe provider metadata without exposing credentials."""
    key_source = ""
    for name in ("LLM_API_KEY", "LLM_KEY", "DEEPSEEK_API_KEY", "DASHSCOPE_API_KEY"):
        if os.getenv(name):
            key_source = name
            break
    return {
        "endpoint": (
            os.getenv("LLM_BASE_URL")
            or os.getenv("LLM_ENDPOINT")
            or os.getenv("DASHSCOPE_BASE_URL")
            or "https://api.deepseek.com/v1"
        ),
        "key_source": key_source,
        "key_present": bool(key_source),
        "key_length": len(os.getenv(key_source, "")) if key_source else 0,
        "model": DEFAULT_FLASH_MODEL,
    }


def _extract_message_text(message):
    if message is None:
        return ""
    if isinstance(message, str):
        return message
    if isinstance(message, list):
        parts = []
        for item in message:
            if isinstance(item, dict):
                part_type = item.get("type")
                if part_type in {"text", "output_text"}:
                    parts.append(item.get("text", ""))
            else:
                parts.append(str(item))
        return "".join(parts)
    return str(message)


def _local_image_to_data_url(image_path: str) -> str:
    path = Path(str(image_path or "").strip())
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"image file not found: {image_path}")
    mime_type, _ = mimetypes.guess_type(str(path))
    normalized_mime_type = str(mime_type or "").strip() or "application/octet-stream"
    raw = path.read_bytes()
    encoded = base64.b64encode(raw).decode("utf-8")
    return f"data:{normalized_mime_type};base64,{encoded}"


def _build_multimodal_content(
    *,
    text: str = "",
    image_urls=None,
    image_paths=None,
    detail: str = "auto",
):
    content = []
    normalized_text = str(text or "").strip()
    if normalized_text:
        content.append({"type": "text", "text": normalized_text})
    for image_url in image_urls or []:
        url = str(image_url or "").strip()
        if not url:
            continue
        content.append({
            "type": "image_url",
            "image_url": {
                "url": url,
                "detail": detail,
            },
        })
    for image_path in image_paths or []:
        path = str(image_path or "").strip()
        if not path:
            continue
        content.append({
            "type": "image_url",
            "image_url": {
                "url": _local_image_to_data_url(path),
                "detail": detail,
            },
        })
    if not content:
        raise ValueError("multimodal content requires text or images")
    return content


def build_multimodal_user_message(
    *,
    text: str = "",
    image_urls=None,
    image_paths=None,
    detail: str = "auto",
):
    return {
        "role": "user",
        "content": _build_multimodal_content(
            text=text,
            image_urls=image_urls,
            image_paths=image_paths,
            detail=detail,
        ),
    }


def _create_llm_completion(
    message_body,
    *,
    model=None,
    max_tokens=None,
    temperature=0.7,
    top_p=0.95,
    response_format=None,
    enable_think=None,
    client_instance=None,
):
    resolved_model = model or DEFAULT_CHAT_MODEL
    request_kwargs = {
        "model": resolved_model,
        "messages": message_body,
        "temperature": temperature,
        "max_tokens": max_tokens or DEFAULT_MAX_TOKENS,
        "top_p": top_p,
        "frequency_penalty": 0,
        "presence_penalty": 0,
        "stop": None,
    }
    if response_format is not None:
        request_kwargs["response_format"] = response_format
    if enable_think is not None:
        request_kwargs["extra_body"] = {"enable_thinking": enable_think}
    return (client_instance or llm_client).chat.completions.create(**request_kwargs)


def create_llm_embeddings(
    texts,
    *,
    model=None,
    dimensions=None,
    client_instance=None,
):
    rows = [str(item or "") for item in (texts or [])]
    if not rows:
        return [], {}
    request_kwargs = {
        "model": model or DEFAULT_EMBEDDING_MODEL,
        "input": rows,
    }
    if dimensions is not None:
        request_kwargs["dimensions"] = int(dimensions)
    response = (client_instance or llm_client).embeddings.create(**request_kwargs)
    vectors = [item.embedding for item in (response.data or [])]
    return vectors, getattr(response, "usage", None)


def _chat_with_default_model(
    message_body,
    *,
    max_tokens=None,
    temperature=0.7,
    top_p=0.95,
    enable_think=None,
):
    response = _create_llm_completion(
        message_body,
        max_tokens=max_tokens,
        temperature=temperature,
        top_p=top_p,
        enable_think=DEFAULT_ENABLE_THINKING if enable_think is None else enable_think,
    )
    return _extract_message_text(response.choices[0].message.content), response.usage


def _chat_json_with_default_model(message_body, *, max_tokens=None, enable_think=None):
    content, usage = _chat_with_default_model(
        message_body,
        max_tokens=max_tokens,
        enable_think=enable_think,
    )
    ret_json = extract_first_json(content)
    if not ret_json:
        print("[ERROR] LLM ret None")
    return ret_json, usage


def _chat_with_flash_model(
    message_body,
    *,
    max_tokens=None,
    temperature=0.7,
    top_p=0.95,
    enable_think=None,
):
    response = _create_llm_completion(
        message_body,
        model=DEFAULT_FLASH_MODEL,
        max_tokens=max_tokens,
        temperature=temperature,
        top_p=top_p,
        enable_think=False if enable_think is None else enable_think,
    )
    return _extract_message_text(response.choices[0].message.content), response.usage


def _chat_json_with_flash_model(
    message_body,
    *,
    max_tokens=None,
    enable_think=None,
    temperature=0.7,
):
    content, usage = _chat_with_flash_model(
        message_body,
        max_tokens=max_tokens,
        temperature=temperature,
        enable_think=enable_think,
    )
    ret_json = extract_first_json(content)
    if not ret_json:
        print("[ERROR] FLASH LLM ret None")
    return ret_json, usage



def extract_first_json(text: str, *, log_errors: bool = True):
    if not isinstance(text, str):
        if log_errors:
            print("not str")
        return None

    # 1) fenced json/object-or-array
    m = re.search(r"```json\s*(.*?)\s*```", text, flags=re.DOTALL | re.IGNORECASE)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            pass

    # 2) first JSON object or array from first opening bracket
    object_start = text.find("{")
    array_start = text.find("[")
    starts = [idx for idx in (object_start, array_start) if idx != -1]
    start = min(starts) if starts else -1
    if start == -1:
        if log_errors:
            print("no tag")
        return None

    depth = 0
    end = None
    opening = text[start]
    closing = "}" if opening == "{" else "]"
    for i, ch in enumerate(text[start:], start=start):
        if ch == opening:
            depth += 1
        elif ch == closing:
            depth -= 1
            if depth == 0:
                end = i
                break

    if end is None:
        if log_errors:
            print("tag not match")
        return None

    try:
        return json.loads(text[start:end + 1])
    except Exception as e:
        if log_errors:
            print("ill format json" , e)
        return None

'''
def extract_first_json(text: str) -> dict:
    if not isinstance(text, str):
        None
    try:
        m = re.search(r"```json\\s*(.*?)\\s*```", text, flags=re.DOTALL | re.IGNORECASE)
        if m:
            print("std pattern match")
            return json.loads(m.group(1))
        start = text.find("{")
        if start == -1:
            print("no json tag")
            return None
        depth = 0
        for i in range(start, len(text)):
            if text[i] == "{":
                depth += 1
            elif text[i] == "}":
                depth -= 1
        if depth == 0:
            return json.loads(text[start:i + 1])
        print("json pattern ill")
        return None
    except Exception as e:
        print("json format error")
        return None
'''


def chat_openai(message_body):
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=message_body,
        temperature=0.7,
        max_tokens=2048,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
    return response.choices[0].message.content

def chat_openai_4omini_json(message_body):
    response = client_next.chat.completions.create(
        model="gpt-4o-mini",
        messages=message_body,
        response_format ={ "type": "json_object" },
        temperature=0.7,
        max_tokens=2048,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
    return response.choices[0].message.content ,response.usage

def chat_openai_4omini(message_body):
    response = client_next.chat.completions.create(
        model="gpt-4o-mini",
        messages=message_body,
        temperature=0.7,
        max_tokens=2048,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
    return response.choices[0].message.content ,response.usage


def chat_qwen(message_body, enable_think=False):
    return _chat_with_default_model(message_body, enable_think=enable_think)


def chat_qwen_long_context(message_body, enable_think=False):
    return _chat_with_default_model(
        message_body,
        max_tokens=DEFAULT_LONG_CONTEXT_MAX_TOKENS,
        enable_think=enable_think,
    )


def chat_qwen_json(message_body, enable_think=False):
    return _chat_json_with_default_model(
        message_body,
        enable_think=enable_think,
    )


def chat_qwen_flash(message_body, enable_think=False):
    return _chat_with_flash_model(message_body, enable_think=enable_think)


def chat_qwen_flash_json(message_body, enable_think=False, temperature=0.7):
    return _chat_json_with_flash_model(
        message_body,
        enable_think=enable_think,
        temperature=temperature,
    )


def chat_qwen_flash_json_with_raw(message_body, enable_think=False, temperature=0.0):
    content, usage = _chat_with_flash_model(
        message_body,
        enable_think=enable_think,
        temperature=temperature,
    )
    return extract_first_json(content, log_errors=False), usage, content


def chat_qwen_flash_structured(message_body, enable_think=False, temperature=0.0):
    response = _create_llm_completion(
        message_body,
        model=DEFAULT_FLASH_MODEL,
        temperature=temperature,
        response_format={"type": "json_object"},
        enable_think=enable_think,
    )
    return _extract_message_text(response.choices[0].message.content), response.usage


def chat_qwen_multimodal(
    *,
    text: str = "",
    image_urls=None,
    image_paths=None,
    message_body=None,
    enable_think=False,
    max_tokens=None,
    temperature=0.7,
    top_p=0.95,
    detail: str = "auto",
):
    messages = message_body
    if messages is None:
        messages = [
            build_multimodal_user_message(
                text=text,
                image_urls=image_urls,
                image_paths=image_paths,
                detail=detail,
            )
        ]
    return _chat_with_default_model(
        messages,
        max_tokens=max_tokens,
        temperature=temperature,
        top_p=top_p,
        enable_think=enable_think,
    )


def chat_qwen_multimodal_json(
    *,
    text: str = "",
    image_urls=None,
    image_paths=None,
    message_body=None,
    enable_think=False,
    max_tokens=None,
    detail: str = "auto",
):
    messages = message_body
    if messages is None:
        messages = [
            build_multimodal_user_message(
                text=text,
                image_urls=image_urls,
                image_paths=image_paths,
                detail=detail,
            )
        ]
    return _chat_json_with_default_model(
        messages,
        max_tokens=max_tokens,
        enable_think=enable_think,
    )


def chat_deepseek(message_body,enable_think=False):
    return chat_qwen(message_body, enable_think=enable_think)


def chat_deepseek_long_context(message_body,enable_think=False):
    return chat_qwen_long_context(message_body, enable_think=enable_think)


# 用普通接口封装json接口，json接口不稳定
def chat_deepseek_json(message_body , enable_think=False):
    return chat_qwen_json(message_body, enable_think=enable_think)


def chat_deepseek_flash(message_body, enable_think=False):
    return chat_qwen_flash(message_body, enable_think=enable_think)


def chat_deepseek_flash_json(message_body, enable_think=False):
    return chat_qwen_flash_json(message_body, enable_think=enable_think)




def excel2md(file, output=None):
    # Open the Excel workbook and evaluate cell formulas
    workbook = openpyxl.load_workbook(file, data_only=True)
    # Select the sheet you want to convert
    sheet = workbook.active

    # Initialize the Markdown table
    markdown_table = ""

    # Determine the range of rows and columns to include in the table
    start_row = 1
    start_column = 1
    end_row = sheet.max_row
    end_column = sheet.max_column

    # Add the table rows
    for row in range(start_row, end_row + 1):
        # Initialize the row string for the Markdown table
        row_string = ""
        # Check if the entire row contains 'None' values
        row_contains_none = all(cell.value is None for cell in
                                [sheet.cell(row=row, column=column) for column in range(start_column, end_column + 1)])

        # Skip the row if it contains only 'None' values
        if row_contains_none:
            continue

        # Add the row to the Markdown table
        for column in range(start_column, end_column + 1):
            # Check if the entire column contains 'None' values
            column_contains_none = all(cell.value is None for cell in
                                       [sheet.cell(row=row_num, column=column) for row_num in
                                        range(start_row, end_row + 1)])

            # Skip the column if it contains only 'None' values
            if column_contains_none:
                continue

            # Add the cell value to the row string
            cell = sheet.cell(row=row, column=column)
            row_string += f"| {cell.value} "

        # Add the row string to the Markdown table
        markdown_table += row_string + "|\n"
    return markdown_table


def chat_deepseek_reasoner(message_body):
    response = _create_llm_completion(
        message_body,
        model=DEFAULT_REASONING_MODEL,
        max_tokens=32000,
        top_p=0.95,
        enable_think=True,
        client_instance=client_alibaba,
    )
    message = response.choices[0].message
    return _extract_message_text(message.content), getattr(message, "reasoning_content", "")

def chat_deepseek_V3(message_body):
    response = _create_llm_completion(
        message_body,
        model=DEFAULT_CHAT_MODEL,
        temperature=0.0001,
        max_tokens=DEFAULT_MAX_TOKENS,
        top_p=0.95,
        enable_think=False,
        client_instance=client_alibaba,
    )
    message = response.choices[0].message
    return _extract_message_text(message.content), getattr(message, "reasoning_content", "")


def chat_deepseek_V3_json(message_body):
    response = _create_llm_completion(
        message_body,
        model=DEFAULT_CHAT_MODEL,
        response_format={
            'type': 'json_object'
        },
        temperature=0.7,
        max_tokens=DEFAULT_MAX_TOKENS,
        top_p=0.95,
        enable_think=False,
        client_instance=client_alibaba,
    )
    message = response.choices[0].message
    return _extract_message_text(message.content), getattr(message, "reasoning_content", "")


def chat_az_4o_mini(message_body):
    # Backward-compatible shim. Existing callers expect a lightweight chat tuple.
    return chat_qwen(message_body, enable_think=False)




def get_azure_document_analysis_client():
    from azure.core.credentials import AzureKeyCredential
    from azure.ai.documentintelligence import DocumentIntelligenceClient

    endpoint = "https://qianxing-doc.cognitiveservices.azure.com/"
    key = "1b479d9c3fad45e79f78f8cdd43cec59"
    document_analysis_client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    return document_analysis_client

'''
def get_embedding(text, model="text-embedding-ada-002"):
    response = az_client.embeddings.create(input=text, model=model)
    embeddings = [response.data[i].embedding for i in range(len(text))]
    # print(len(response))
    return embeddings
'''
# test the function

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Local AI service smoke runner")
    parser.add_argument("--mode", default="chat", choices=["chat", "interaction_preprocess"])
    parser.add_argument("--text", default="What is the capital of France?")
    parser.add_argument("--work-context-json", default="{}")
    parser.add_argument("--enable-think", action="store_true")
    args = parser.parse_args()

    if args.mode == "interaction_preprocess":
        from src.prompting.prompt_registry import get_prompt_registry

        try:
            work_context = json.loads(str(args.work_context_json or "{}"))
        except Exception:
            work_context = {}
        registry = get_prompt_registry()
        message_body = registry.render_messages(
            "system.assistant.interaction_preprocess",
            {
                "user_text": str(args.text or ""),
                "work_context": work_context if isinstance(work_context, dict) else {},
            },
        )
        payload, usage = chat_qwen_json(message_body, enable_think=bool(args.enable_think))
        print(json.dumps({"payload": payload, "usage": {
            "prompt_tokens": int(getattr(usage, "prompt_tokens", 0) or 0),
            "completion_tokens": int(getattr(usage, "completion_tokens", 0) or 0),
            "total_tokens": int(getattr(usage, "total_tokens", 0) or 0),
        }}, ensure_ascii=False, indent=2))
    else:
        message_body = [
            {
                "role": "user",
                "content": str(args.text or "What is the capital of France?")
            }
        ]
        response, _ = chat_deepseek(message_body, enable_think=bool(args.enable_think))
        print(response)
