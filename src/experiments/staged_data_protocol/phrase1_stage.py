from __future__ import annotations

import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Mapping, Sequence

from openpyxl import Workbook

from src.experiments.staged_data_protocol.phase2.catalog import (
    CATALOG_PATH,
    load_catalog_source,
)

DEFAULT_PROMPT_PATH = Path("phrase_1_prompt.md")
DEFAULT_CASES_PATH = Path("src/experiments/staged_data_protocol/fixtures/cases.json")
DEFAULT_STAGE1_CATALOG_PATH = CATALOG_PATH


def load_questions(path: str | Path | None = None) -> List[Dict[str, str]]:
    case_path = Path(path) if path else DEFAULT_CASES_PATH
    payload = json.loads(case_path.read_text(encoding="utf-8"))
    items = _case_items(payload)
    rows: List[Dict[str, str]] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, Mapping):
            continue
        question = _question_from_item(item)
        if not question:
            continue
        rows.append(
            {
                "case_id": str(item.get("case_id") or f"case_{index:03d}").strip(),
                "question": question,
            }
        )
    return rows


def run_phrase1_cases(
    questions: Sequence[Mapping[str, str]],
    *,
    prompt_path: str | Path = DEFAULT_PROMPT_PATH,
    workers: int = 1,
) -> Dict[str, Any]:
    prompt_file = Path(prompt_path)
    prompt_template = render_stage1_prompt_template(prompt_file.read_text(encoding="utf-8"))
    contract = parse_subject_dataview_contract(prompt_template)
    rows = _run_cases_concurrently(
        questions,
        prompt_template=prompt_template,
        contract=contract,
        workers=workers,
    )
    return {
        "experiment": "phrase1_stage1",
        "prompt_path": str(prompt_file),
        "case_count": len(rows),
        "workers": max(1, int(workers or 1)),
        "contract": contract,
        "summary": summarize(rows),
        "cases": rows,
    }


def render_stage1_prompt_template(
    prompt_template: str,
    *,
    catalog_path: str | Path = DEFAULT_STAGE1_CATALOG_PATH,
) -> str:
    if "{{subject_dataview_context}}" not in prompt_template:
        return prompt_template
    return prompt_template.replace(
        "{{subject_dataview_context}}",
        build_subject_dataview_context(catalog_path=catalog_path),
    )


def build_subject_dataview_context(
    *,
    catalog_path: str | Path = DEFAULT_STAGE1_CATALOG_PATH,
) -> str:
    catalog_file = Path(catalog_path)
    payload = (
        load_catalog_source()
        if catalog_file.resolve() == CATALOG_PATH.resolve()
        else json.loads(catalog_file.read_text(encoding="utf-8"))
    )
    subjects = payload.get("subjects")
    if not isinstance(subjects, Mapping):
        return ""

    lines: List[str] = []
    for subject, subject_cfg in subjects.items():
        if str(subject).startswith("_") or not isinstance(subject_cfg, Mapping):
            continue
        meta = subject_cfg.get("_meta")
        subject_desc = _one_line_desc(meta.get("desc")) if isinstance(meta, Mapping) else ""
        lines.append(f"- {subject}({subject_desc})" if subject_desc else f"- {subject}")
        for dataview, view_cfg in subject_cfg.items():
            if str(dataview).startswith("_") or not isinstance(view_cfg, Mapping):
                continue
            route_summary = _one_line_desc(view_cfg.get("desc"))
            lines.append(
                f"  + {dataview}: {route_summary}"
                if route_summary
                else f"  + {dataview}:"
            )
    return "\n".join(lines)


def _one_line_desc(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _run_cases_concurrently(
    questions: Sequence[Mapping[str, str]],
    *,
    prompt_template: str,
    contract: Mapping[str, Sequence[str]],
    workers: int,
) -> List[Dict[str, Any]]:
    worker_count = max(1, int(workers or 1))
    if worker_count == 1:
        return [
            run_phrase1_case(
                row,
                prompt_template=prompt_template,
                contract=contract,
            )
            for row in questions
        ]
    results: List[Dict[str, Any] | None] = [None] * len(questions)
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(
                run_phrase1_case,
                row,
                prompt_template=prompt_template,
                contract=contract,
            ): index
            for index, row in enumerate(questions)
        }
        for future in as_completed(futures):
            index = futures[future]
            try:
                results[index] = future.result()
            except Exception as exc:  # noqa: BLE001 - keep batch output inspectable.
                row = questions[index]
                results[index] = {
                    "case_id": str(row.get("case_id") or "").strip(),
                    "question": str(row.get("question") or "").strip(),
                    "analyze": "",
                    "steps": None,
                    "parsed_steps": [],
                    "validation_errors": [f"runtime_error: {exc}"],
                    "judgement": f"需修正：runtime_error: {exc}",
                    "raw_response": "",
                    "parse_error": "",
                    "raw_payload": {},
                }
    return [row for row in results if row is not None]


def run_phrase1_case(
    row: Mapping[str, str],
    *,
    prompt_template: str,
    contract: Mapping[str, Sequence[str]],
) -> Dict[str, Any]:
    question = str(row.get("question") or "").strip()
    prompt = prompt_template.replace("{{question}}", question)
    llm_result = call_llm_json(prompt)
    payload = llm_result["payload"]
    parsed_steps = parse_step_rows(payload.get("steps") if isinstance(payload, Mapping) else None)
    validation_errors = validate_phrase1_output(payload, parsed_steps, contract)
    if llm_result.get("parse_error"):
        validation_errors.insert(0, f"parse_error: {llm_result['parse_error']}")
    return {
        "case_id": str(row.get("case_id") or "").strip(),
        "question": question,
        "analyze": str(payload.get("analyze") or "") if isinstance(payload, Mapping) else "",
        "steps": payload.get("steps") if isinstance(payload, Mapping) else None,
        "parsed_steps": parsed_steps,
        "validation_errors": validation_errors,
        "judgement": "结构通过" if not validation_errors else "需修正：" + "；".join(validation_errors[:3]),
        "raw_response": llm_result.get("raw_response", ""),
        "parse_error": llm_result.get("parse_error", ""),
        "raw_payload": payload,
    }


def parse_subject_dataview_contract(prompt: str) -> Dict[str, List[str]]:
    contract: Dict[str, List[str]] = {}
    current_subject = ""
    for line in prompt.splitlines():
        subject_match = re.match(r"^\s*-\s*([A-Za-z0-9_]+)\s*(?:[（(]|$)", line)
        if subject_match:
            current_subject = subject_match.group(1).strip()
            contract.setdefault(current_subject, [])
            continue
        view_match = re.match(r"^\s*\+\s*([A-Za-z0-9_ ]+)\s*[:：]", line)
        if current_subject and view_match:
            dataview = view_match.group(1).strip()
            if dataview:
                contract[current_subject].append(dataview)
    return contract


def _case_items(payload: Any) -> List[Any]:
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, Mapping):
        return []
    for key in ["cases", "rows", "test_cases", "items"]:
        value = payload.get(key)
        if isinstance(value, list):
            return value
    return []


def _question_from_item(item: Mapping[str, Any]) -> str:
    for key in ["question", "query", "user_question"]:
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    input_value = item.get("input")
    if isinstance(input_value, str) and input_value.strip():
        return input_value.strip()
    if isinstance(input_value, Mapping):
        messages = input_value.get("messages")
        if isinstance(messages, list):
            for message in reversed(messages):
                if not isinstance(message, Mapping):
                    continue
                if message.get("role") == "user" and str(message.get("content") or "").strip():
                    return str(message.get("content") or "").strip()
    return ""


def parse_step_rows(steps: Any) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    if not isinstance(steps, list):
        return rows
    for index, item in enumerate(steps, start=1):
        text = str(item or "").strip()
        parts = [part.strip() for part in text.split("|")]
        row = {
            "raw": text,
            "step_id": parts[0] if len(parts) > 0 else "",
            "subject": parts[1] if len(parts) > 1 else "",
            "dataview": parts[2] if len(parts) > 2 else "",
            "condition_desc": parts[3] if len(parts) > 3 else "",
            "part_count": str(len(parts)),
            "row_index": str(index),
        }
        rows.append(row)
    return rows


def validate_phrase1_output(
    payload: Any,
    parsed_steps: Sequence[Mapping[str, str]],
    contract: Mapping[str, Sequence[str]],
) -> List[str]:
    errors: List[str] = []
    if not isinstance(payload, Mapping):
        return ["LLM output is not a JSON object"]
    if not str(payload.get("analyze") or "").strip():
        errors.append("missing analyze")
    if not isinstance(payload.get("steps"), list) or not payload.get("steps"):
        errors.append("steps must be a non-empty list")
    seen_steps: set[str] = set()
    for step in parsed_steps:
        step_id = str(step.get("step_id") or "").strip()
        subject = str(step.get("subject") or "").strip()
        dataview = str(step.get("dataview") or "").strip()
        part_count = str(step.get("part_count") or "")
        prefix = step_id or f"row_{step.get('row_index')}"
        if part_count != "4":
            errors.append(f"{prefix}: step must have 4 pipe parts")
        if not re.match(r"^S\d+$", step_id):
            errors.append(f"{prefix}: invalid step id")
        if step_id in seen_steps:
            errors.append(f"{prefix}: duplicate step id")
        seen_steps.add(step_id)
        if subject not in contract:
            errors.append(f"{prefix}: unknown subject {subject}")
            continue
        if dataview not in set(contract.get(subject) or []):
            errors.append(f"{prefix}: unknown dataview {subject}.{dataview}")
    return errors


def summarize(rows: Sequence[Mapping[str, Any]]) -> Dict[str, int]:
    valid_cases = sum(1 for row in rows if not row.get("validation_errors"))
    step_count = sum(len(row.get("parsed_steps") or []) for row in rows)
    error_count = sum(len(row.get("validation_errors") or []) for row in rows)
    return {
        "valid_cases": valid_cases,
        "case_count": len(rows),
        "step_count": step_count,
        "error_count": error_count,
    }


def write_excel(payload: Mapping[str, Any], output_path: str | Path) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "cases"
    ws.append(
        [
            "case_id",
            "question",
            "judgement",
            "validation_errors",
            "analyze",
            "steps",
            "parsed_steps_json",
            "parse_error",
            "raw_response",
        ]
    )
    for row in payload.get("cases") or []:
        ws.append(
            [
                row.get("case_id"),
                row.get("question"),
                row.get("judgement"),
                "\n".join(row.get("validation_errors") or []),
                row.get("analyze"),
                "\n".join(str(item) for item in (row.get("steps") or [])),
                json.dumps(row.get("parsed_steps") or [], ensure_ascii=False),
                row.get("parse_error"),
                row.get("raw_response"),
            ]
        )
    step_ws = wb.create_sheet("steps")
    step_ws.append(["case_id", "step_id", "subject", "dataview", "condition_desc", "raw"])
    for row in payload.get("cases") or []:
        for step in row.get("parsed_steps") or []:
            step_ws.append(
                [
                    row.get("case_id"),
                    step.get("step_id"),
                    step.get("subject"),
                    step.get("dataview"),
                    step.get("condition_desc"),
                    step.get("raw"),
                ]
            )
    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["metric", "value"])
    for key, value in (payload.get("summary") or {}).items():
        summary_ws.append([key, value])
    wb.save(path)


def call_llm_json(prompt: str) -> Dict[str, Any]:
    from src.utils.ai_service import chat_qwen_flash

    raw_response, _usage = chat_qwen_flash([{"role": "user", "content": prompt}], enable_think=False)
    try:
        payload = json.loads(_extract_json_text(raw_response))
        if not isinstance(payload, dict):
            return {"payload": {}, "raw_response": raw_response, "parse_error": "top-level JSON is not an object"}
        return {"payload": payload, "raw_response": raw_response, "parse_error": ""}
    except Exception as exc:  # noqa: BLE001 - experiment should preserve model parse failures.
        return {"payload": {}, "raw_response": raw_response, "parse_error": str(exc)}


def _extract_json_text(text: str) -> str:
    raw = str(text or "").strip()
    fenced = re.search(r"```(?:json)?\s*(.*?)```", raw, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        return fenced.group(1).strip()
    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        return raw[start : end + 1]
    return raw
